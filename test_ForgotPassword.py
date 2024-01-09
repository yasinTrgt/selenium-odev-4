from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalConstants as c


class Test_ForgotPassword:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.FORGOT_PASSWORD_URL)
        self.driver.maximize_window()
        sleep(3)  # import time

    def teardown_method(self): # her testinin bitiminde çalışacak fonk
           self.driver.quit()    

    def getData():
        excel = openpyxl.load_workbook(c.LOGIN_TOBETO_XLSX)
        sheet = excel["Sheet4"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(1,rows+1):
            forgotPassword = sheet.cell(i,2).value
            data.append((forgotPassword))

        return data

           
           
    @pytest.mark.parametrize("email",getData())
    def forgot_password(self,forgotPassword):
         forgotPasswordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.FORGOT_PASSWORD)))
         forgotPasswordInput.send_keys(forgotPassword)
         gonder = self.driver.find_element(By.XPATH, "//*[@id='__next']/div/main/section/div/div/div/button")
         gonder.click()
         sleep(3)

